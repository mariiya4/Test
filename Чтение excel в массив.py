#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import numpy

def sheet_to_array(filepath, sheetname=None):
    WB=openpyxl.load_workbook(filepath)
    if sheetname is None:
        sheetname=WB.sheetnames[0]  #если лист не указан то первый
    sh=WB[sheetname]
    
    kol_strok=sh.max_row-sh.min_row+1
    kol_stolb=sh.max_column-sh.min_column+1
    Result=np.zeros( (kol_strok,kol_stolb),dtype=np.float64)
    
    for kol_strok, row in enumerate(sh.rows):
        for kol_stolb,cell in enumerate(row):
            Result[kol_strok,kol_stolb]=cell.value
            
    return Result